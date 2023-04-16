import locale
import os
from datetime import datetime, timedelta

import xlrd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.styles import (
    Border, Alignment, Font, Side, PatternFill,
    NamedStyle
)


common = NamedStyle(name="common")
common.font = Font("Times New Roman", size=12)
bd = Side(style="thin", color="000000")
common.border = Border(left=bd, top=bd, right=bd, bottom=bd)
common.alignment = Alignment(horizontal="center")

common_bold = NamedStyle(name="common_bold")
common_bold.font = Font("Times New Roman", size=12, bold=True)
bd = Side(style="thin", color="000000")
common_bold.border = Border(left=bd, top=bd, right=bd, bottom=bd)
common_bold.alignment = Alignment(horizontal="center")
common_bold.number_format = "0"

grey = NamedStyle(name='grey')
grey.font = Font("Times New Roman", size=12)
bd = Side(style="thin", color="00000000")
grey.border = Border(left=bd, top=bd, right=bd, bottom=bd)
grey.alignment = Alignment(horizontal="center")
grey.fill = PatternFill(fill_type="solid", start_color="bfbfbf", end_color="bfbfbf")

orange = NamedStyle(name='orange')
orange.font = Font("Times New Roman", size=12, bold=True, color="ff0000")
bd = Side(style="thin", color="00000000")
orange.border = Border(left=bd, top=bd, right=bd, bottom=bd)
orange.alignment = Alignment(horizontal="center")
orange.fill = PatternFill(fill_type="solid", start_color="ffc000",
                     end_color="ffc000")
orange.number_format = '0'

common_blue = NamedStyle(name='common_blue')
common_blue.font = Font("Times New Roman", size=12, bold=True, color="0070c0")
bd = Side(style="thin", color="00000000")
common_blue.border = Border(left=bd, top=bd, right=bd, bottom=bd)
common_blue.alignment = Alignment(horizontal="center")
common_blue.number_format = '0'

common_green = NamedStyle(name='common_green')
common_green.font = Font("Times New Roman", size=12, bold=True, color="00b050")
bd = Side(style="thin", color="00000000")
common_green.border = Border(left=bd, top=bd, right=bd, bottom=bd)
common_green.alignment = Alignment(horizontal="center")
common_green.number_format = '0'

styles = (common, grey, orange, common_blue, common_green, common_bold)


def get_data_from(path):
    book = xlrd.open_workbook(path)
    filepath, filename = os.path.split(path)
    if str(filename).__contains__("УПВСН Миннибаево"):
        table = {
            't_input': [5, 20],
            't_output': [6, 20],
            'p_input': [7, 20],
            'p_output': [8, 20],
            't_flue_gas': [9, 20],
            'q_oil_daily': [10, 21],
            'water_cut': [11, 20],
            'salt_content': [12, 20],
            'q_gas_daily': [14, 21],
            'gas_calorific_value': [15, 8],
            'efficiency': [16, 8]
        }
    else:
        table = {
            't_input': [4, 20],
            't_output': [5, 20],
            'p_input': [6, 20],
            'p_output': [7, 20],
            'q_oil_daily': [9, 21],
            'water_cut': [10, 20],
            'salt_content': [11, 20],
            'q_gas_daily': [13, 21],
            'gas_calorific_value': [14, 8],
            'efficiency': [15, 8]
        }
    data = {}
    sheet_count = 1 + book.sheet_by_index(-1).number
    for sheet in range(sheet_count):
        for col_name, coord in table.items():
            value = book.sheet_by_index(sheet).cell_value(
                        colx=coord[0], rowx=coord[1])
            if isinstance(value, float):
                if sheet+1 not in data:
                    data[sheet+1] = {col_name:round(value,2)}
                else:
                    data[sheet + 1].update({col_name: round(value, 2)})
            else:
                if sheet+1 not in data:
                    data[sheet + 1] = {col_name: None}
                else:
                    data[sheet + 1].update({col_name: None})

    return data


def write_data_to(data, sheet_name):
    book = load_workbook(
        "D:\\Users\\obs\\Downloads\\ПЕЧИ ЯН новый формат v1.2.003.xlsx"
        )
    for style in styles:
        try:
            book.add_named_style(style)
        except ValueError:
            pass
    if sheet_name == "ГЗНУ-4304":
        sheet_name = "ГЗНУ-4304(ЦДНГ-3)"
    sheet = book[sheet_name]
    start_row = sheet.max_row
    start_date = sheet[f'A{start_row}'].value
    row = start_row
    date = start_date
    for keys, values in data.items():
        row += 1
        date += timedelta(days=1)
        t_input = values['t_input']
        t_output = values['t_output']
        t_delta = f"=C{row}-B{row}"
        q_oil_daily = values['q_oil_daily']
        water_cut = values['water_cut']
        salt_content = values['salt_content']
        gas_calorific_value = values['gas_calorific_value']
        t_environment = None
        q_gas_daily = values['q_gas_daily']
        q_gas_delta = f"=L{row}-M{row}"
        q_gas_growing = f"=N{row}+O{row-1}"
        bot_line = f"=N{row}-1.96*2984.66"
        top_line = f"=N{row}+1.96*2984.66"
        global q_gas_calc, efficiency
        efficiency = f"=(G{row}*0.5+G{row}/(100-H{row})*H{row})*(D{row})/" \
                     f"(L{row}*J{row}/1000)*100"
        if sheet_name == "УПВСН":
            q_gas_delta = f"=M{row}-N{row}"
            q_gas_growing = f"=O{row}+P{row - 1}"
            q_gas_calc = f"=1059.889757*1-183.5582496*C{row}-0.759521028*" \
                         f"L{row}+2.26034617*H{row}+192.5080014*D{row}"
            efficiency = f"=IF(B{row}=0,0,(H{row}*0.5+H{row}/(100-I{row})*" \
                         f"I{row})*D{row}/(M{row}*K{row}/1000)*100)"
        elif sheet_name == "ДНС-1":
            q_gas_calc = f"=1062.712205*1-47.8252062*C{row}-1.72087028*" \
                         f"K{row}+0.070152858*G{row}+92.56899132*D{row}"
        elif sheet_name == "ДНС-210":
            q_gas_calc = f"=1646.230614*1-5.801667948*C{row}+" \
                         f"1.593840077*K{row}-0.039029513*G{row}+" \
                         f"6.033148611*D{row}"
        elif sheet_name == "ДНС-8":
            q_gas_calc = f"=501.4371862*1+131.0892944*C{row}+" \
                         f"8.642538027*K{row}+0.807459557*G{row}-" \
                         f"150.8270611*D{row}"
        elif sheet_name == "ГЗНУ-4304(ЦДНГ-3)":
            q_gas_calc = f"=1376.880844*1-73.99592473*C{row}+15.99007925*" \
                         f"K{row}+0.012856337*G{row}+85.37297771*D{row}"
        if sheet_name == "ДНС-1" or sheet_name == "ДНС-210":
            try:
                p_input = values['p_input'] / 10
                p_output = values['p_output'] / 10
            except TypeError:
                p_input = p_output = None
        else:
            p_input = values['p_input']
            p_output = values['p_output']

        date_cell = sheet.cell(row=row, column=1, value=date)
        date_cell.style = "common"
        date_cell.number_format='DD/MM/YYYY'
        sheet.cell(row=row, column=2, value=t_input).style = "common"
        sheet.cell(row=row, column=3, value=t_output).style = "common"
        sheet.cell(row=row, column=4, value=t_delta).style = "common"
        sheet.cell(row=row, column=5, value=p_input).style = "common"
        sheet.cell(row=row, column=6, value=p_output).style = "common"

        if sheet_name == "УПВСН":
            t_flue_gas = values['t_flue_gas']
            sheet.cell(row=row, column=7, value=t_flue_gas).style = "common"
            sheet.cell(row=row, column=8, value=q_oil_daily).style = "common"
            sheet.cell(row=row, column=9, value=water_cut).style = "common"
            sheet.cell(row=row, column=10, value=salt_content).style = "common"
            sheet.cell(row=row, column=11, value=gas_calorific_value).style = "common"
            sheet.cell(row=row, column=12, value=t_environment).style = "common"
            sheet.cell(row=row, column=13, value=q_gas_daily).style = "grey"
            sheet.cell(row=row, column=14, value=q_gas_calc).style = "orange"
            sheet.cell(row=row, column=15, value=q_gas_delta).style = "common_blue"
            sheet.cell(row=row, column=16, value=q_gas_growing).style = "common_green"
            sheet.cell(row=row, column=17, value=bot_line).style = "common"
            sheet.cell(row=row, column=18, value=top_line).style = "common"
            sheet.cell(row=row, column=19, value=efficiency).style = "common_bold"

        else:
            sheet.cell(row=row, column=7, value=q_oil_daily).style = "common"
            sheet.cell(row=row, column=8, value=water_cut).style = "common"
            sheet.cell(row=row, column=9, value=salt_content).style = "common"
            sheet.cell(row=row, column=10, value=gas_calorific_value).style = "common"
            sheet.cell(row=row, column=11, value=t_environment).style = "common"
            sheet.cell(row=row, column=12, value=q_gas_daily).style = "grey"
            sheet.cell(row=row, column=13, value=q_gas_calc).style = "orange"
            sheet.cell(row=row, column=14, value=q_gas_delta).style = "common_blue"
            sheet.cell(row=row, column=15, value=q_gas_growing).style = "common_green"
            sheet.cell(row=row, column=16, value=efficiency).style =  "common_bold"

        last_row = sheet.max_row
    for column in range(1, sheet.max_column):
        sheet.cell(row=start_row, column=column).border = Border(
            bottom=Side(border_style="medium", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000")
        )
        sheet.cell(row=last_row, column=column).border = Border(
            bottom=Side(border_style="medium", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000")
        )


    lc = LineChart()
    lc.legend.position = "b"
    anchor = TwoCellAnchor()
    anchor._from.col = 16
    anchor._from.row = start_row
    anchor.to.col = 35
    anchor.to.row = last_row
    locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
    lc.title = datetime.strftime(start_date+timedelta(days=1), u'%B %Y')
    date_values = Reference(sheet, min_col=1, max_col=1, min_row=start_row+1,
                     max_row=last_row)
    if sheet_name == "УПВСН":
        anchor._from.col = 19
        anchor.to.col = 38
        q_gas_values = Reference(sheet, min_col=13, max_col=13,
                                 min_row=start_row + 1, max_row=last_row)
        q_gas_calc_values = Reference(sheet, min_col=14, max_col=14,
                                      min_row=start_row, max_row=last_row)
        bot_line_values = Reference(sheet, min_col=17, max_col=17,
                                    min_row=start_row, max_row=last_row)
        top_line_values = Reference(sheet, min_col=18, max_col=18,
                                    min_row=start_row + 1, max_row=last_row)
        bot_line_series = Series(values=bot_line_values, title="Нижняя граница")
        top_line_series = Series(values=top_line_values, title="Верхняя граница")
        lc.append(bot_line_series)
        lc.append(top_line_series)
    else:
        q_gas_values = Reference(sheet, min_col=12, max_col=12,
                                 min_row=start_row+1, max_row=last_row)
        q_gas_calc_values = Reference(sheet, min_col=13, max_col=13,
                                 min_row=start_row, max_row=last_row)
    q_gas_series = Series(values=q_gas_values, title="Qг, н.м3/сут")
    q_gas_calc_series = Series(values=q_gas_calc_values, title="Qг, РАСЧЕТ н.м3/сут ")
    lc.append(q_gas_series)
    lc.append(q_gas_calc_series)
    lc.set_categories(date_values)
    sheet.add_chart(lc, f"Q{start_row+1}")
    lc.anchor = anchor

    book.save('D:\\Users\\obs\\Downloads\\ПЕЧИ ЯН новый формат v1.2.003.xlsx')



